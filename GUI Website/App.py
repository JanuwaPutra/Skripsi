from flask import Flask, render_template, request
import cv2
import numpy as np
import pandas as pd
from skimage.feature import graycomatrix, graycoprops
from flask import Flask, url_for
from openpyxl import load_workbook


app = Flask(__name__)

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/deteksi')
def deteksi():
    return render_template('deteksi.html')

@app.route('/predict', methods=['POST'])
def predict():
    # Memuat data
        file        = '../PROGRAM/basisdatafitur2.xlsx'
        dataset     = pd.read_excel(file)
        glcm_properties = ['dissimilarity', 'correlation', 'homogeneity', 'contrast',  'energy']

        fitur       = dataset.iloc[:,+1:-1].values
        kelas       = dataset.iloc[:,24].values
        tes_fitur   = []
        tes_fitur.append([])
        file_name = request.files['file'].filename  # Mengambil nama file
         
        # Feature extraction for data testing----------------------------------------------
        # Preprocessing
        src = cv2.imdecode(np.frombuffer(request.files['file'].read(), np.uint8), cv2.IMREAD_COLOR)
        src     = cv2.resize(src, (640, 640))

        #segmentasi
        tmp     = cv2.cvtColor(src, cv2.COLOR_BGR2GRAY)
        _,mask  = cv2.threshold(tmp,140,255,cv2.THRESH_BINARY_INV)
        mask    = cv2.dilate(mask.copy(),None,iterations=20)
        mask    = cv2.erode(mask.copy(),None,iterations=25)
        b, g, r = cv2.split(src)
        rgba    = [b,g,r, mask]
        dst     = cv2.merge(rgba,4)

        contours, hierarchy = cv2.findContours(mask,cv2.RETR_TREE,cv2.CHAIN_APPROX_SIMPLE)
        selected    = max(contours,key=cv2.contourArea)
        x,y,w,h     = cv2.boundingRect(selected)
        cropped     = dst[y:y+h,x:x+w]
        mask        = mask[y:y+h,x:x+w]
        gray        = cv2.cvtColor(cropped, cv2.COLOR_BGR2GRAY)


        # RGB-HSV
        hsv_image   = cv2.cvtColor(cropped, cv2.COLOR_BGR2HSV)
           
        H = hsv_image[:,:,0]
        S = hsv_image[:,:,1]
        V = hsv_image[:,:,2]
                    
        H = np.mean(H)
        S = np.mean(S)
        V = np.mean(V)

        tes_fitur[0].append(H)
        tes_fitur[0].append(S)
        tes_fitur[0].append(V)

        # GLCM

        # Calculate GLCM with desired distances
        glcm = graycomatrix(gray, distances=[5], angles=[0, np.pi/4, np.pi/2, 3*np.pi/4],
                   levels=256, symmetric=True, normed=True)
        feature = []
        glcm_props = [propery for name in glcm_properties for propery in graycoprops(glcm, name)[0]]
        for item in glcm_props:            
            tes_fitur[0].append(item)


        # --------------------------------------------------
        # Menghitung mean dan standar deviasi dari data fitur
        mean = np.mean(fitur, axis=0)
        std_dev = np.std(fitur, axis=0)

        # Menstandarisasi data fitur dan data uji menggunakan mean dan standar deviasi
        fitur = (fitur - mean) / std_dev
        tes_fitur_std = (tes_fitur - mean) / std_dev

        # Hitung jarak euclidean 
        distances = np.sqrt(np.sum(np.square(fitur - tes_fitur_std), axis=1))

        # Sortir indeks berdasarkan jarak dan ambil k terdekat
        nearest_indices = np.argsort(distances)[:3]

        # Ambil kelas dari indeks terdekat
        nearest_classes = kelas[nearest_indices]

        # Hitung kelas yang paling sering muncul
        unique_classes, counts = np.unique(nearest_classes, return_counts=True)
        y_pred = unique_classes[np.argmax(counts)]

        if 'segar' in y_pred:
            #untuk hasil output web
            hasil = 'Ikan Segar'
            #untuk hasil memasukkan database fitur
            labeldeep = 'ikansegar'
        else:
            hasil = 'Ikan Busuk'
            labeldeep = 'ikanbusuk'
            
        tambah_fitur_baru(file,file_name, tes_fitur[0], labeldeep )
    
        return render_template('result.html', prediction=hasil)
    
def tambah_fitur_baru(file, file_name, fitur_baru, labeldeep):
    """
    Menambah fitur baru ke file Excel.
    """
    # Baca workbook dan pilih sheet pertama
    wb = load_workbook(file)
    sheet = wb.active
    
    # Tentukan baris terakhir yang terisi
    last_row = sheet.max_row
    
    # Tentukan kolom untuk menambah data baru
    col_offset = 1
    
    # Tambahkan nama file ke kolom pertama
    sheet.cell(row=last_row + 1, column=col_offset, value=file_name)
    
    # Tambahkan fitur baru ke kolom berikutnya
    for i, value in enumerate(fitur_baru):
        sheet.cell(row=last_row + 1, column=col_offset + 1 + i, value=value)
    
    # Tambahkan label ke kolom terakhir
    sheet.cell(row=last_row + 1, column=col_offset + len(fitur_baru) + 1, value=labeldeep)
    
    # Simpan workbook
    wb.save(file)
    wb.close()


    
if __name__ == '__main__':
    app.run(debug=True)